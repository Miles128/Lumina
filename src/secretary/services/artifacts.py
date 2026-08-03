"""Browse / preview local workspace & thread-sandbox files for the chat artifact panel."""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any

from secretary.agent import thread_sandbox
from secretary.services.agent_config import AgentConfigStore

MAX_PREVIEW_BYTES = 2 * 1024 * 1024
MAX_RAW_BYTES = 25 * 1024 * 1024
MAX_TREE_ENTRIES = 400
MAX_TREE_DEPTH = 6
TEXT_EXTS = frozenset(
    {
        ".md",
        ".markdown",
        ".txt",
        ".json",
        ".csv",
        ".tsv",
        ".yaml",
        ".yml",
        ".toml",
        ".py",
        ".js",
        ".ts",
        ".xml",
        ".log",
        ".sh",
        ".rs",
        ".go",
        ".java",
        ".c",
        ".h",
        ".cpp",
        ".swift",
        ".kt",
        ".sql",
        ".ini",
        ".cfg",
        ".conf",
        ".env",
        ".gitignore",
        ".dockerfile",
    }
)
HTML_EXTS = frozenset({".html", ".htm"})
TABLE_EXTS = frozenset({".xlsx", ".xlsm"})
PDF_EXTS = frozenset({".pdf"})
DOCX_EXTS = frozenset({".docx"})


def _resolve_existing(raw: str | Path) -> Path:
    path = Path(raw).expanduser()
    try:
        return path.resolve()
    except OSError:
        return path.absolute()


def allowed_roots(*, data_dir: Path, thread_id: str, agent_config: AgentConfigStore) -> list[dict[str, str]]:
    roots: list[dict[str, str]] = []
    sandbox = thread_sandbox.ensure(thread_id, data_dir)
    roots.append({"id": "sandbox", "label": "会话沙箱", "path": str(sandbox)})
    cwd = agent_config.load().shell_working_dir.strip()
    if cwd:
        path = _resolve_existing(cwd)
        if path.is_dir():
            roots.append({"id": "workspace", "label": "工作区", "path": str(path)})
    return roots


def _is_under(path: Path, root: Path) -> bool:
    try:
        return path.is_relative_to(root)
    except (ValueError, OSError):
        return False


def assert_path_allowed(
    raw: str,
    *,
    data_dir: Path,
    thread_id: str,
    agent_config: AgentConfigStore,
    tree_only: bool = False,
) -> Path:
    target = _resolve_existing(raw)
    roots = [_resolve_existing(item["path"]) for item in allowed_roots(
        data_dir=data_dir, thread_id=thread_id, agent_config=agent_config
    )]
    if any(_is_under(target, root) for root in roots):
        return target
    if tree_only:
        raise PermissionError(f"path outside allowed roots: {target}")
    # Local-first desktop: allow preview of agent-written files under home / data dir.
    home = Path.home().resolve()
    data = _resolve_existing(data_dir)
    if _is_under(target, home) or _is_under(target, data):
        return target
    raise PermissionError(f"path outside allowed roots: {target}")


def list_tree(
    raw: str,
    *,
    data_dir: Path,
    thread_id: str,
    agent_config: AgentConfigStore,
    depth: int = 3,
) -> dict[str, Any]:
    root = assert_path_allowed(
        raw,
        data_dir=data_dir,
        thread_id=thread_id,
        agent_config=agent_config,
        tree_only=True,
    )
    if not root.exists():
        raise FileNotFoundError(str(root))
    if not root.is_dir():
        raise NotADirectoryError(str(root))
    max_depth = max(1, min(int(depth or 3), MAX_TREE_DEPTH))
    entries: list[dict[str, Any]] = []
    skipped = 0

    def walk(current: Path, level: int) -> None:
        nonlocal skipped
        if len(entries) >= MAX_TREE_ENTRIES:
            skipped += 1
            return
        try:
            children = sorted(current.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
        except OSError:
            return
        for child in children:
            if len(entries) >= MAX_TREE_ENTRIES:
                skipped += 1
                break
            name = child.name
            if name.startswith(".") and name not in {".env", ".gitignore"}:
                continue
            try:
                is_dir = child.is_dir()
            except OSError:
                continue
            rel = str(child.relative_to(root))
            item: dict[str, Any] = {
                "name": name,
                "path": str(child),
                "rel": rel,
                "type": "dir" if is_dir else "file",
                "depth": level,
            }
            if not is_dir:
                item["ext"] = child.suffix.lower()
                try:
                    item["size"] = child.stat().st_size
                except OSError:
                    item["size"] = 0
            entries.append(item)
            if is_dir and level + 1 < max_depth:
                walk(child, level + 1)

    walk(root, 0)
    return {
        "root": str(root),
        "entries": entries,
        "truncated": skipped > 0,
        "count": len(entries),
    }


def _read_text_preview(path: Path) -> str:
    data = path.read_bytes()
    if len(data) > MAX_PREVIEW_BYTES:
        data = data[:MAX_PREVIEW_BYTES]
    return data.decode("utf-8", errors="replace")


def _csv_table(text: str, *, max_rows: int = 80) -> dict[str, Any]:
    reader = csv.reader(io.StringIO(text))
    rows: list[list[str]] = []
    truncated = False
    for idx, row in enumerate(reader):
        if idx >= max_rows:
            truncated = True
            break
        rows.append([str(cell) for cell in row])
    return {"rows": rows, "truncated": truncated}


def _xlsx_table(path: Path, *, max_rows: int = 80) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            truncated = False
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx >= max_rows:
                    truncated = True
                    break
                cells = [("" if cell is None else str(cell)) for cell in row]
                if any(cells):
                    rows.append(cells)
            sheets.append({"name": ws.title, "rows": rows, "truncated": truncated})
        return {"sheets": sheets}
    finally:
        wb.close()


def resolve_raw_file(
    raw: str,
    *,
    data_dir: Path,
    thread_id: str,
    agent_config: AgentConfigStore,
) -> Path:
    path = assert_path_allowed(raw, data_dir=data_dir, thread_id=thread_id, agent_config=agent_config)
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(str(path))
    size = path.stat().st_size
    if size > MAX_RAW_BYTES:
        raise ValueError(f"file too large (max {MAX_RAW_BYTES // (1024 * 1024)} MB)")
    return path


def preview_file(
    raw: str,
    *,
    data_dir: Path,
    thread_id: str,
    agent_config: AgentConfigStore,
) -> dict[str, Any]:
    path = assert_path_allowed(raw, data_dir=data_dir, thread_id=thread_id, agent_config=agent_config)
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(str(path))
    size = path.stat().st_size
    ext = path.suffix.lower()
    result: dict[str, Any] = {
        "path": str(path),
        "name": path.name,
        "ext": ext,
        "size": size,
        "kind": "binary",
    }

    if ext in TABLE_EXTS:
        if size > MAX_PREVIEW_BYTES:
            raise ValueError(f"file too large for preview (max {MAX_PREVIEW_BYTES // (1024 * 1024)} MB)")
        result["kind"] = "table"
        result["table"] = _xlsx_table(path)
        return result

    if ext in PDF_EXTS:
        result["kind"] = "pdf"
        if size <= MAX_PREVIEW_BYTES:
            try:
                from secretary.agent.tools.documents import _read_pdf

                result["text"] = _read_pdf(path, page_start=1, page_end=8)
            except Exception as exc:  # noqa: BLE001 — preview must not fail embed
                result["text"] = ""
                result["note"] = f"text extract unavailable: {exc}"
        else:
            result["text"] = ""
            result["note"] = "large PDF — embed only"
        return result

    if ext in DOCX_EXTS:
        if size > MAX_PREVIEW_BYTES:
            raise ValueError(f"file too large for preview (max {MAX_PREVIEW_BYTES // (1024 * 1024)} MB)")
        from secretary.agent.tools.documents import _read_docx

        result["kind"] = "docx"
        result["text"] = _read_docx(path, max_paragraphs=160)
        return result

    if ext in HTML_EXTS:
        if size > MAX_PREVIEW_BYTES:
            raise ValueError(f"file too large for preview (max {MAX_PREVIEW_BYTES // (1024 * 1024)} MB)")
        result["kind"] = "html"
        result["text"] = _read_text_preview(path)
        return result

    if size > MAX_PREVIEW_BYTES:
        raise ValueError(f"file too large for preview (max {MAX_PREVIEW_BYTES // (1024 * 1024)} MB)")

    if ext in TEXT_EXTS or ext == "" or path.name in {"Makefile", "Dockerfile", "LICENSE", "README"}:
        text = _read_text_preview(path)
        result["kind"] = "markdown" if ext in {".md", ".markdown"} else "text"
        result["text"] = text
        if ext == ".csv":
            result["kind"] = "table"
            result["table"] = {"sheets": [{"name": path.name, **_csv_table(text)}]}
        elif ext == ".json":
            try:
                result["text"] = json.dumps(json.loads(text), ensure_ascii=False, indent=2)
            except json.JSONDecodeError:
                pass
        return result
    # Fallback: try utf-8 text
    try:
        sample = path.read_bytes()[:4096]
        sample.decode("utf-8")
        result["kind"] = "text"
        result["text"] = _read_text_preview(path)
        return result
    except UnicodeDecodeError as exc:
        raise ValueError(f"unsupported binary type: {ext or path.name}") from exc
